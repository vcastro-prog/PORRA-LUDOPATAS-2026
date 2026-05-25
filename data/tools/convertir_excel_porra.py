from __future__ import annotations

from pathlib import Path
import re
import pandas as pd
from openpyxl import load_workbook


def _clean(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def convertir_excel_porra(ruta_excel: str | Path, carpeta_salida: str | Path = "data") -> tuple[pd.DataFrame, pd.DataFrame]:
    """Convierte el Excel original de la porra a tablas limpias.

    El formato esperado es:
    - una hoja por participante
    - nombre del participante en B3
    - partidos en bloques con columnas: fecha, PARTIDO n, local, goles local, visitante, goles visitante
    - los bloques comienzan en columnas C, I y O

    Devuelve:
    - apuestas: participante, partido_id, goles_local, goles_visitante
    - partidos: partido_id, grupo, fecha, local, visitante
    """
    ruta_excel = Path(ruta_excel)
    carpeta_salida = Path(carpeta_salida)
    carpeta_salida.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(ruta_excel, data_only=True)
    apuestas_rows: list[dict] = []
    partidos_por_id: dict[int, dict] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        participante = _clean(ws["B3"].value) or _clean(sheet_name)
        grupo_actual = ""

        for row in range(1, ws.max_row + 1):
            posible_grupo = _clean(ws.cell(row, 1).value)
            if posible_grupo.startswith("Gp."):
                grupo_actual = posible_grupo

            for col_partido in (3, 9, 15):
                etiqueta = _clean(ws.cell(row, col_partido).value)
                match = re.search(r"PARTIDO\s*(\d+)", etiqueta.upper())
                if not match:
                    continue

                partido_id = int(match.group(1))
                fecha = _clean(ws.cell(row, col_partido - 1).value)
                local = _clean(ws.cell(row, col_partido + 1).value)
                visitante = _clean(ws.cell(row, col_partido + 3).value)
                gl = ws.cell(row, col_partido + 2).value
                gv = ws.cell(row, col_partido + 4).value

                if local and visitante:
                    partidos_por_id.setdefault(
                        partido_id,
                        {
                            "partido_id": partido_id,
                            "grupo": grupo_actual,
                            "fecha": fecha,
                            "local": local,
                            "visitante": visitante,
                        },
                    )

                try:
                    gl = int(gl)
                    gv = int(gv)
                except Exception:
                    continue

                apuestas_rows.append(
                    {
                        "participante": participante,
                        "partido_id": partido_id,
                        "goles_local": gl,
                        "goles_visitante": gv,
                    }
                )

    apuestas = pd.DataFrame(apuestas_rows).sort_values(["participante", "partido_id"])
    partidos = pd.DataFrame(partidos_por_id.values()).sort_values("partido_id")

    apuestas.to_csv(carpeta_salida / "apuestas_reales.csv", index=False)
    partidos.to_csv(carpeta_salida / "partidos.csv", index=False)

    resultados = partidos[["partido_id"]].copy()
    resultados["goles_local"] = pd.NA
    resultados["goles_visitante"] = pd.NA
    resultados.to_csv(carpeta_salida / "resultados.csv", index=False)

    return apuestas, partidos


if __name__ == "__main__":
    convertir_excel_porra("data/apuestas_base.xlsx", "data")
    print("Generados data/apuestas_reales.csv, data/partidos.csv y data/resultados.csv")
