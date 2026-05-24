from __future__ import annotations

from pathlib import Path
from typing import BinaryIO

import pandas as pd
from openpyxl import load_workbook


def _to_int_or_none(value):
    if value is None or value == "":
        return None
    try:
        return int(value)
    except Exception:
        return None


def leer_apuesta_excel(file: str | Path | BinaryIO, partidos: pd.DataFrame) -> pd.DataFrame:
    wb = load_workbook(file, data_only=True)
    ws = wb.active
    participante = ws["B3"].value or "Sin nombre"
    participante = str(participante).strip()
    rows = []
    for _, p in partidos.iterrows():
        gl = _to_int_or_none(ws.cell(int(p["row"]), int(p["home_goals_col"])).value)
        gv = _to_int_or_none(ws.cell(int(p["row"]), int(p["away_goals_col"])).value)
        if gl is None or gv is None:
            continue
        rows.append({"participante": participante, "partido_id": int(p["partido_id"]), "goles_local": gl, "goles_visitante": gv})
    return pd.DataFrame(rows)
